import openpyxl
import datetime
import os

def createNewSheet():
    today = datetime.date.today()
    date = today.strftime('%d-%m-%Y')

    workbook = openpyxl.Workbook()
    
    directory = 'sheet/'
    if not os.path.exists(directory):
        os.mkdir(directory)


    workbook.save(directory + date +'.xlsx')
    


def writeToSheet(r, c, count):
    today = datetime.date.today()
    date = today.strftime('%d-%m-%Y')

    directory = 'sheet/'

    if not os.path.exists(directory + date + '.xlsx'):
        createNewSheet()

    workbook = openpyxl.load_workbook(directory + date +'.xlsx')
    worksheet = workbook.active

    cell = worksheet.cell(row=r, column=c)
    cell.value = count

    workbook.save(directory + date +'.xlsx')