from openpyxl.reader.excel import ExcelReader
from openpyxl.xml import constants as openpyxl_xml_constants
from pandas import ExcelFile
from pandas.io.excel._openpyxl import OpenpyxlReader
from openpyxl.packaging.relationship import (
    RelationshipList,
    get_dependents,
    get_rels_path,
)
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet._reader import WorksheetReader
from openpyxl.worksheet.table import Table
from openpyxl.xml.constants import (
    COMMENTS_NS
)
from openpyxl.comments.comment_sheet import CommentSheet
from openpyxl.cell import MergedCell
from openpyxl.comments.comment_sheet import CommentSheet
from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing

from openpyxl.xml.functions import fromstring

from openpyxl.pivot.table import TableDefinition

from openpyxl.xml.functions import fromstring


class CustomExcelReader(ExcelReader):
    def read_worksheets(self):
        for sheet, rel in self.parser.find_sheets():
            if rel.target not in self.valid_files:
                continue

            if "chartsheet" in rel.Type:
                self.read_chartsheet(sheet, rel)
                continue

            rels_path = get_rels_path(rel.target)
            rels = RelationshipList()
            if rels_path in self.valid_files:
                rels = get_dependents(self.archive, rels_path)

            if self.read_only:
                ws = ReadOnlyWorksheet(self.wb, sheet.name, rel.target, self.shared_strings)
                ws.sheet_state = sheet.state
                self.wb._sheets.append(ws)
                continue
            else:
                fh = self.archive.open(rel.target)
                ws = self.wb.create_sheet(sheet.name)
                ws._rels = rels
                ws_parser = WorksheetReader(ws, fh, self.shared_strings, self.data_only, self.rich_text)
                ws_parser.bind_all()

            # preserve link to VML file if VBA
            if self.wb.vba_archive and ws.legacy_drawing:
                ws.legacy_drawing = rels[ws.legacy_drawing].target
            else:
                ws.legacy_drawing = None

            for t in ws_parser.tables:
                src = self.archive.read(t)
                xml = fromstring(src)
                table = Table.from_tree(xml)
                ws.add_table(table)

            pivot_rel = rels.find(TableDefinition.rel_type)
            for r in pivot_rel:
                pivot_path = r.Target
                src = self.archive.read(pivot_path)
                tree = fromstring(src)
                pivot = TableDefinition.from_tree(tree)
                pivot.cache = self.parser.pivot_caches[pivot.cacheId]
                ws.add_pivot(pivot)

            ws.sheet_state = sheet.state

class OpenpyxlReaderWOFormatting(OpenpyxlReader):

    def load_workbook(self, filepath_or_buffer):
        reader = CustomExcelReader(filepath_or_buffer, read_only=False, data_only=True, keep_links=True)
        reader.archive.read = self.read_exclude_styles(reader.archive)
        reader.read()
        return reader.wb

    def read_exclude_styles(self, archive):
        """skips addings styles to xlsx workbook , like they were absent
        see logic in openpyxl.styles.stylesheet.apply_stylesheet """

        orig_read = archive.read

        def new_read(name, pwd=None):
            if name == openpyxl_xml_constants.ARC_STYLE:
                raise KeyError
            else:
                return orig_read(name, pwd=pwd)

        return new_read



def monkey_patch_openpyxl():
    '''Openpyxl has a bug with workbooks that have wrong cell styling information.
    Monkey patch the library so it can handle these types of workbooks.'''
    from openpyxl.worksheet import _reader
    from openpyxl.cell import Cell
    def bind_cells(self):
        for idx, row in self.parser.parse():
            for cell in row:
                try:
                    style = self.ws.parent._cell_styles[cell['style_id']]
                except:  ## This is the patch, original doesn't have a try/except here
                    style = None
                c = Cell(self.ws, row=cell['row'], column=cell['column'], style_array=style)
                c._value = cell['value']
                c.data_type = cell['data_type']
                self.ws._cells[(cell['row'], cell['column'])] = c
        #self.ws.formula_attributes = self.parser.array_formula
        if self.ws._cells:
            self.ws._current_row = self.ws.max_row # use cells not row dimensions

    _reader.WorksheetReader.bind_cells = bind_cells


    def bind_formatting(self):
        x = 1
        
    _reader.WorksheetReader.bind_formatting = bind_formatting

    def bind_col_dimensions(self):
        x = 1

    _reader.WorksheetReader.bind_col_dimensions = bind_col_dimensions

    def bind_row_dimensions(self):
        x = 1

    _reader.WorksheetReader.bind_row_dimensions = bind_row_dimensions


monkey_patch_openpyxl()


ExcelFile._engines['openpyxl_wo_formatting'] = OpenpyxlReaderWOFormatting

