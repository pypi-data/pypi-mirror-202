import xlrd2
import pyexcel
import openpyxl

from ka_ut_com.log import Log

from ka_ut_obj.arrs import Arrs
from ka_ut_obj.arr import Arr


class Xlrd2:

    @staticmethod
    def read_workbook(path, index=0):
        """ read Workbook sheet into Array of Arrays
        """
        workbook = xlrd2.open_workbook(path)
        sheet = workbook.sheet_by_index(index)
        aoa = []
        for ii in range(1, sheet.nrows):
            row = sheet.row_values(ii)
            aoa.append(row)
        return aoa


class PyExcel_:

    @staticmethod
    def read_workbook(path):
        """ read Workbook sheet into Array of Arrays
        """
        workbook = pyexcel.get_book(file_name=path)
        Log.Eq.debug("workbook", workbook)
        sheet = workbook[0]
        Log.Eq.debug("sheet", sheet)
        aoa = []
        for ii in range(1, sheet.number_of_rows()):
            row = sheet.row(ii)
            aoa.append(row)
        return aoa


class Openpyxl_:

    @staticmethod
    def load_workbook(path, **kwargs):
        print("==============================")
        print(f"path = {path}")
        print("==============================")
        return openpyxl.load_workbook(path, **kwargs)

    @staticmethod
    def get_workbook(**kwargs):
        return openpyxl.Workbook(**kwargs)


class Xlsx:

    class AoA:

        @classmethod
        def read_workbook(path, sheet_id):
            """ read Workbook sheet into Array of Arrays
            """
            # return cls.PyExcel.read_workbook(path, index)
            workbook = Workbook.load(path)

            Log.Eq.debug("sheet_id", sheet_id)
            if isinstance(sheet_id, int):
                sheet_names = workbook.sheetnames
                sheet = workbook[sheet_names[sheet_id]]
                Log.Eq.debug("sheet_names", sheet_names)
            else:
                sheet = workbook[sheet_id]
            Log.Eq.debug("workbook", workbook)
            rows = list(sheet.values)[1:]
            Log.Eq.debug("rows", rows)
            aoa = []
            for row in rows:
                if row[2] is None:
                    continue
                aoa.append(row)
            return aoa


class Sheet:

    class Headers:

        @staticmethod
        def iter_column(sheet, **kwargs):
            row_header = kwargs.get('headers_start', 1)
            col_name_prefix = kwargs.get('col_name_prefix', '')
            col_start = 1
            col_end = sheet.max_column+1
            sw_add_sheet_name = kwargs.get('sw_add_sheet_name', False)
            if sw_add_sheet_name:
                header_sheet_name = kwargs.get(
                                      'header_sheet_name', 'sheet_name')
                col_name = f"{col_name_prefix}{header_sheet_name}"
                yield col_name
            for col in range(col_start, col_end):
                col_name = sheet.cell(column=col, row=row_header).value
                col_name = f"{col_name_prefix}{col_name}"
                yield col_name

    @staticmethod
    def set(sheet, rows):
        for row in rows:
            sheet.append(row)

    @staticmethod
    def iter_row(row, **kwargs):
        sw_add_sheet_name = kwargs.get('sw_add_sheet_name', False)
        if sw_add_sheet_name:
            sheet_name = kwargs.get('sheet_name')
            yield sheet_name
        for cell in row:
            yield cell.value

    @staticmethod
    def iter_rows(sheet):
        for row in sheet.iter_rows():
            yield [cell.value for cell in row]

    @classmethod
    def iter_sheet_lst(cls, sheet, **kwargs):
        row_start = kwargs.get('row_start', 1)
        row_end = kwargs.get('row_end', sheet.max_row)
        col_start = kwargs.get('col_start', 1)
        col_end = kwargs.get('col_end', sheet.max_column)

        for row in sheet.iter_rows(
                     min_row=row_start, min_col=col_start,
                     max_row=row_end, max_col=col_end):
            yield list(cls.iter_row(row, **kwargs))

    @classmethod
    def sh_headers(cls, sheet, **kwargs):
        return list(cls.Headers.iter_column(sheet, **kwargs))

    @classmethod
    def sh_aoa(cls, sheet, **kwargs):
        return list(cls.iter_sheet_lst(sheet, **kwargs))

    @staticmethod
    def sh_id(sheet_name=None, sheet_index=0):
        if sheet_name is not None:
            return sheet_name
        return sheet_index

    @classmethod
    def to_aod(cls, sheet):
        rows = list(cls.iter_rows(sheet))
        headers = rows[0]
        Log.Eq.debug("headers", headers)
        aod = []
        for row in rows[1:]:
            dic = Arrs.sh_dic(headers, Arr.apply_str(row))
            aod.append(dic)
        Log.Eq.debug("aod[0]", aod[0])
        return aod

    @classmethod
    def to_dic(cls, sheet):
        rows = list(cls.iter_rows(sheet))
        headers = rows[0]
        dic = {}
        if len(headers) == 2:
            for row in rows[1:]:
                dic[row[0]] = row[1]
        else:
            for row in rows[1:]:
                dic[row[0]] = row[1]
        Log.Eq.debug("dic", dic)
        return dic

    @classmethod
    def to_arr(cls, sheet):
        rows = list(cls.iter_rows(sheet))
        # headers = rows[0]
        arr = []
        for row in rows[1:]:
            if row[0] is not None and row[1] is not None:
                arr.append(row)
        return arr


class Workbook:

    @staticmethod
    def load(path):
        return Openpyxl_.load_workbook(path, read_only=True)

    @staticmethod
    def get():
        return Openpyxl_.get_workbook(write_only=True)

    @staticmethod
    def iter_sheet_names(workbook, **kwargs):
        cols_count = kwargs.get('cols_count')
        sheet_names = kwargs.get('sheet_names')
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            if sheet.max_column == cols_count:
                yield sheet_name

    @classmethod
    def sh_sheet_names(cls, workbook, **kwargs):
        cols_count = kwargs.get('cols_count')
        sheet_names = kwargs.get('sheet_names')
        if sheet_names is None:
            sheet_names = workbook.sheetnames
        if cols_count is None:
            return sheet_names
        return list(cls.iter_sheet_names(
                      workbook,
                      cols_count=cols_count,
                      sheet_names=sheet_names))

    @classmethod
    def read_aoa(cls, path, **kwargs):
        workbook = cls.load(path)
        heads_sheet_name = kwargs.get('headers_sheet_name')
        aoa = []
        sheet_names = cls.sh_sheet_names(workbook, **kwargs)
        if heads_sheet_name is not None:
            sheet = workbook[heads_sheet_name]
            heads = Sheet.sh_headers(sheet, **kwargs)
        else:
            heads = []
        for sheet_name in sheet_names:
            Log.Eq.debug("sheet_name", sheet_name)
            sheet = workbook[sheet_name]
            aoa_sheet = Sheet.sh_aoa(
                          sheet,
                          sheet_name=sheet_name,
                          **kwargs)
            aoa.extend(aoa_sheet)
            Log.Eq.debug("aoa_sheet", aoa_sheet)
        return heads, aoa

    @classmethod
    def ex_read_aoa(cls, **kwargs):
        prefix = kwargs.get('prefix')
        in_path = kwargs.get(f'in_path_{prefix}')
        row_start = kwargs.get(f'row_start_{prefix}')
        cols_count = kwargs.get(f'cols_count_{prefix}')
        sw_add_sheet_name = kwargs.get(f'sw_add_sheet_name_{prefix}')
        sheet_names = kwargs.get(f'sheet_names_{prefix}')
        headers_sheet_name = kwargs.get(f'headers_sheet_name_{prefix}')
        headers_start = kwargs.get(f'headers_start_{prefix}')

        return cls.read_aoa(
                 in_path,
                 row_start=row_start,
                 cols_count=cols_count,
                 sw_add_sheet_name=sw_add_sheet_name,
                 sheet_names=sheet_names,
                 headers_sheet_name=headers_sheet_name,
                 headers_start=headers_start)

    @staticmethod
    def iter_sheet(workbook, max_sheets):
        for ii in range(0, max_sheets):
            yield workbook.create_sheet()

    @classmethod
    def write_dod_sheet(cls, path, dod_sheet, **kwargs):
        workbook = cls.get()
        for name in dod_sheet.keys():
            rows = dod_sheet[name]['rows']
            sheet = workbook.create_sheet()
            sheet.title = name
            Sheet.set(sheet, rows)
        workbook.save(path)

    @staticmethod
    def sh_sheet(workbook, sheet_id):
        if isinstance(sheet_id, int):
            sheet_names = workbook.sheetnames
            sheet = workbook[sheet_names[sheet_id]]
        else:
            sheet = workbook[sheet_id]
        return sheet

    @classmethod
    def read_sheet_to_dic(cls, path, sheet_id):
        workbook = Workbook.load(path)
        sheet = cls.sh_sheet(workbook, sheet_id)
        return Sheet.to_dic(sheet)

    @classmethod
    def read_sheet_to_arr(cls, path, sheet_id):
        workbook = Workbook.load(path)
        sheet = cls.sh_sheet(workbook, sheet_id)
        return Sheet.to_arr(sheet)

    @classmethod
    def read_sheets_to_aoa(cls, path, sheet_ids=None):
        workbook = Workbook.load(path)
        if sheet_ids is None:
            sheet_ids = workbook.sheetnames
        for sheet_id in sheet_ids:
            yield cls.read_sheet_to_arr(path, sheet_id)

    @staticmethod
    def read_sheet_names(path):
        workbook = Workbook.load(path)
        return workbook.sheetnames

    @classmethod
    def read_workbooks_sheets_to_ao_aoa(cls, paths, sheet_ids=None):
        for path in paths:
            yield cls.read_sheets_to_aoa(path, sheet_ids=None)
